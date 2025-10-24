"""
Blueprint for GPA calculator API routes.
"""

from flask import Blueprint, jsonify, request, send_file
import json
import os
import sqlite3
import openpyxl
from datetime import datetime

gpa_bp = Blueprint('api', __name__)

def get_db_connection():
    """Create a database connection."""
    conn = sqlite3.connect(os.path.join('data', 'user_gpa.db'))
    conn.row_factory = sqlite3.Row
    return conn

@gpa_bp.route('/subjects', methods=['GET'])
def get_subjects():
    """Get list of subjects from JSON file."""
    try:
        with open(os.path.join('data', 'subjects.json'), 'r', encoding='utf-8') as f:
            data = json.load(f)
            return jsonify(data['subjects'])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@gpa_bp.route('/grades', methods=['GET', 'POST'])
def handle_grades():
    """Handle grade operations."""
    if request.method == 'POST':
        data = request.json
        
        # Validate input
        required_fields = ['semester', 'subject', 'credits', 'grade10', 'grade4']
        if not all(field in data for field in required_fields):
            return jsonify({'error': 'Missing required fields'}), 400
            
        try:
            conn = get_db_connection()
            c = conn.cursor()
            
            c.execute('''INSERT INTO grades 
                        (user_id, semester, subject, credits, grade_10, grade_4)
                        VALUES (?, ?, ?, ?, ?, ?)''',
                     (1, data['semester'], data['subject'], 
                      data['credits'], data['grade10'], data['grade4']))
            
            conn.commit()
            return jsonify({'success': True, 'id': c.lastrowid})
            
        except Exception as e:
            return jsonify({'error': str(e)}), 500
        finally:
            conn.close()
            
    else:  # GET request
        try:
            user_id = request.args.get('user_id', 1)
            conn = get_db_connection()
            c = conn.cursor()
            
            grades = c.execute('''SELECT * FROM grades 
                                WHERE user_id = ? 
                                ORDER BY semester, subject''',
                             (user_id,)).fetchall()
            
            result = [{
                'id': grade['id'],
                'semester': grade['semester'],
                'subject': grade['subject'],
                'credits': grade['credits'],
                'grade10': grade['grade_10'],
                'grade4': grade['grade_4']
            } for grade in grades]
            
            return jsonify(result)
            
        except Exception as e:
            return jsonify({'error': str(e)}), 500
        finally:
            conn.close()

@gpa_bp.route('/export-excel', methods=['GET'])
def export_excel():
    """Export grades to Excel file."""
    try:
        user_id = request.args.get('user_id', 1)
        conn = get_db_connection()
        c = conn.cursor()
        
        # Get user info
        user = c.execute('SELECT * FROM users WHERE id = ?', 
                        (user_id,)).fetchone()
        
        # Get grades
        grades = c.execute('''SELECT * FROM grades 
                            WHERE user_id = ? 
                            ORDER BY semester, subject''',
                         (user_id,)).fetchall()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bảng điểm"
        
        # Add header
        ws['A1'] = "TRƯỜNG ĐẠI HỌC CÔNG NGHỆ TPHCM"
        ws['A2'] = "BẢNG ĐIỂM SINH VIÊN"
        ws['A3'] = f"Họ và tên: {user['full_name'] if user else ''}"
        ws['A4'] = f"MSSV: {user['username'] if user else ''}"
        ws['A5'] = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y')}"
        
        # Add column headers
        headers = ["Kỳ", "Môn học", "Tín chỉ", "Điểm 10", "Điểm 4"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=7, column=col, value=header)
        
        # Add grades
        row = 8
        current_semester = None
        semester_credits = 0
        semester_points = 0
        
        for grade in grades:
            if current_semester != grade['semester']:
                if current_semester is not None:
                    # Add semester summary
                    semester_gpa = (semester_points / semester_credits 
                                  if semester_credits > 0 else 0)
                    ws.cell(row=row, column=1, 
                           value=f"GPA Kỳ {current_semester}")
                    ws.cell(row=row, column=3, 
                           value=semester_credits)
                    ws.cell(row=row, column=5, 
                           value=round(semester_gpa, 2))
                    row += 2
                
                current_semester = grade['semester']
                semester_credits = 0
                semester_points = 0
            
            ws.cell(row=row, column=1, value=grade['semester'])
            ws.cell(row=row, column=2, value=grade['subject'])
            ws.cell(row=row, column=3, value=grade['credits'])
            ws.cell(row=row, column=4, value=grade['grade_10'])
            ws.cell(row=row, column=5, value=grade['grade_4'])
            
            semester_credits += grade['credits']
            semester_points += grade['credits'] * grade['grade_4']
            row += 1
        
        # Add final semester summary
        if current_semester is not None:
            semester_gpa = (semester_points / semester_credits 
                          if semester_credits > 0 else 0)
            ws.cell(row=row, column=1, 
                   value=f"GPA Kỳ {current_semester}")
            ws.cell(row=row, column=3, 
                   value=semester_credits)
            ws.cell(row=row, column=5, 
                   value=round(semester_gpa, 2))
            row += 2
        
        # Add total GPA
        total_credits = sum(grade['credits'] for grade in grades)
        total_points = sum(grade['credits'] * grade['grade_4'] 
                         for grade in grades)
        total_gpa = total_points / total_credits if total_credits > 0 else 0
        
        ws.cell(row=row, column=1, value="GPA TOÀN KHÓA")
        ws.cell(row=row, column=3, value=total_credits)
        ws.cell(row=row, column=5, value=round(total_gpa, 2))
        
        # Style the worksheet
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Header style
        for row in range(1, 6):
            ws['A' + str(row)].font = Font(bold=True)
        
        # Column headers style
        header_fill = PatternFill(start_color='004D40', 
                                end_color='004D40', 
                                fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col in range(1, 6):
            cell = ws.cell(row=7, column=col)
            cell.fill = header_fill
            cell.font = header_font
        
        # Save file
        filename = f"GPA_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('data', filename)
        wb.save(filepath)
        
        return send_file(filepath, 
                        as_attachment=True,
                        download_name=filename)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()
        if 'filepath' in locals():
            os.remove(filepath)  # Clean up temporary file

@gpa_bp.route('/sync-portal', methods=['POST'])
def sync_portal():
    """Sync grades from UTH portal."""
    # TODO: Implement portal synchronization
    return jsonify({'message': 'Portal sync not implemented yet'}), 501